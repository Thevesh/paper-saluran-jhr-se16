"""
Assemble the final deliverable: one workbook, plus the headline table on stdout.

The headline estimate is the reconciled (raked) one from 03. Two intervals are carried
alongside it, and they mean genuinely different things -- conflating them is the easiest way
to overstate what this exercise can support:

  BOUNDS  what arithmetic alone guarantees. Assumption-free. The truth is inside this
          interval with certainty, not with 95% confidence.
  CI      cluster bootstrap over duns, around the reconciled estimate. This is sampling
          uncertainty in the seed *given* the model's structure. It says nothing about
          whether that structure is right, so it is a floor on uncertainty, never the total.

`model_range` is the spread of the point estimate across M1-M5. It is the most honest quick
read on specification sensitivity: where it is wide, the answer is being chosen by the
assumption rather than found in the data, whatever the CI says.

Read a row as trustworthy when the bounds are narrow AND the model range is small. Where the
bounds are near-vacuous (Indian, Other) the estimate is an extrapolation and should be
reported as such or not at all.
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import common
from common import DIMS, ESTIMAND_NAMES

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")


def sheet(wb, name, df, note=None):
    ws = wb.create_sheet(name)
    if note:
        ws.append([note])
        ws.cell(row=1, column=1).font = Font(italic=True, color="555555")
        ws.append([])
    start = ws.max_row + 1 if note else 1
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for j, c in enumerate(df.columns, start=1):
        cell = ws.cell(row=start, column=j)
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, Alignment(horizontal="center")
        width = max([len(str(c))] + [len(str(v)) for v in df[c].head(300)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(width + 2, 10), 40)
    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    return ws


def run(el):
    OUT = common.outdir(el)
    rec = pd.read_csv(f"{OUT}/reconciled.csv")
    models = pd.read_csv(f"{OUT}/model_estimates.csv")
    bounds = pd.read_csv(f"{OUT}/bounds_statewide.csv")

    spread = (models[models.model != "M1_lpm"]
              .groupby(["dim", "group", "estimand"]).estimate.agg(["min", "max"])
              .rename(columns={"min": "model_min", "max": "model_max"}).reset_index())
    final = rec.merge(spread, on=["dim", "group", "estimand"])
    final["model_range"] = final.model_max - final.model_min
    final["bound_width"] = final.bound_hi - final.bound_lo

    # A flag the reader can act on, rather than a number they have to interpret.
    final["reliability"] = np.where(
        final.bound_width > 0.5, "weak - bounds near-vacuous, driven by assumption",
        np.where(final.model_range > 0.05, "moderate - specification-sensitive", "good"))

    order = {d: i for i, d in enumerate(DIMS)}
    gorder = {g: i for d in DIMS.values() for i, g in enumerate(d)}
    eorder = {e: i for i, e in enumerate(ESTIMAND_NAMES)}
    final = final.sort_values(
        by=["dim", "group", "estimand"],
        key=lambda s: s.map({**order, **gorder, **eorder}) if s.name != "dim" else s.map(order))

    cols = ["dim", "group", "estimand", "registered", "reconciled", "ci_lo", "ci_hi",
            "bound_lo", "bound_hi", "bound_width", "model_min", "model_max", "model_range",
            "capped", "m2_raw", "reliability"]
    final = final[cols]
    final.to_csv(f"{OUT}/FINAL_estimates.csv", index=False)

    print("\n" + "=" * 100)
    print(f"{common.cfg(el)['label']} -- ECOLOGICAL ESTIMATES  (all figures as % of the group's "
          "REGISTERED voters)")
    print("Denominator is registration, so abstention is a modelled choice: "
          "'support' = share of the whole group.")
    print("=" * 100)
    for dim, gs in DIMS.items():
        print(f"\n### by {dim.upper()}\n")
        print(f"{'group':9s} {'registered':>10s} {'estimand':9s} {'EST':>7s} "
              f"{'95% CI':>15s} {'bounds':>15s} {'models':>15s}  reliability")
        for g in gs:
            for e in ESTIMAND_NAMES:
                r = final[(final.dim == dim) & (final.group == g) &
                          (final.estimand == e)].iloc[0]
                print(f"{g:9s} {r.registered:10,} {e:9s} {100 * r.reconciled:6.1f}% "
                      f"{100 * r.ci_lo:6.1f}-{100 * r.ci_hi:<6.1f} "
                      f"{100 * r.bound_lo:6.1f}-{100 * r.bound_hi:<6.1f} "
                      f"{100 * r.model_min:6.1f}-{100 * r.model_max:<6.1f}  {r.reliability}")
            print()

    wb = Workbook()
    wb.remove(wb.active)
    pc = lambda d, cs: d.assign(**{c: (100 * d[c]).round(2) for c in cs})
    sheet(wb, "ESTIMATES", pc(final, [c for c in cols if final[c].dtype.kind == "f"]),
          note="Johor SE-16 group estimates, % of the group's REGISTERED voters. "
               "'bounds' hold with certainty; 'CI' is sampling error only, conditional on "
               "the model being right; 'model_range' is specification sensitivity.")
    sheet(wb, "MODELS", pc(models, ["estimate", "se", "ci_lo", "ci_hi"]),
          note="Every model, unconstrained. M1 LPM is free to leave [0,1] and does so; "
               "M2/M3 are the same model under different links and agree to ~0.01pp.")
    sheet(wb, "BOUNDS", pc(bounds, ["lo", "hi", "width"]),
          note="Deterministic Frechet bounds. Assumption-free; every estimate must lie inside.")
    try:
        v = pd.read_csv(f"{OUT}/validation_sims.csv")
        agg = v.groupby(["dim", "world", "estimand"]).agg(
            bias_pp=("err_recon", lambda x: round(100 * x.mean(), 2)),
            mae_pp=("err_recon", lambda x: round(100 * x.abs().mean(), 2)),
            bounds_cover=("covered", "mean")).reset_index()
        sheet(wb, "VALIDATION", agg,
              note="Recovery of a planted truth on real Johor geography. World A: model "
                   "assumption true. World B: aggregation bias planted. World C: dun "
                   "heterogeneity. World B is the one that matters.")
    except FileNotFoundError:
        print("\n(no validation_sims.csv -- run validate.py to add the VALIDATION sheet)")

    out = f"{OUT}/{el}_estimates.xlsx"
    wb.save(out)
    print(f"\nwrote {OUT}/FINAL_estimates.csv and {out}")


def main():
    for el in common.arg_elections():
        run(el)


if __name__ == "__main__":
    main()
